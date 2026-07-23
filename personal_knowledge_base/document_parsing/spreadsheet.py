import io
import posixpath
import xml.etree.ElementTree as ET
from datetime import date, datetime, time
from zipfile import BadZipFile, ZipFile

from .types import ParsedDocument, TextBlock


class SpreadsheetParseError(ValueError):
    def __init__(self, code: str, message: str):
        self.code = code
        super().__init__(f"{code}: {message}")


_WORKBOOK_NAMESPACE = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_DOCUMENT_RELATIONSHIPS_NAMESPACE = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_PACKAGE_RELATIONSHIPS_NAMESPACE = "http://schemas.openxmlformats.org/package/2006/relationships"
_WORKBOOK_PART = "xl/workbook.xml"
_WORKBOOK_RELATIONSHIPS_PART = "xl/_rels/workbook.xml.rels"


def _safe_xlsx_part(base_part: str, target: str) -> str | None:
    """Resolve an internal OOXML target without allowing archive traversal."""
    if not target or "#" in target or "\\" in target:
        return None
    if target.startswith("/"):
        candidate = target.lstrip("/")
    else:
        candidate = posixpath.join(posixpath.dirname(base_part), target)
    normalized = posixpath.normpath(candidate)
    if normalized.startswith("../") or normalized in {".", ".."} or not normalized.startswith("xl/"):
        return None
    return normalized


def _xlsx_xml(payload: bytes) -> ET.Element:
    if b"<!DOCTYPE" in payload.upper() or b"<!ENTITY" in payload.upper():
        raise SpreadsheetParseError("xlsx_parse_failed", "XLSX XML declarations are not supported")
    return ET.fromstring(payload)


def _xlsx_merged_ranges(data: bytes) -> dict[str, list[str]]:
    """Read worksheet merge metadata directly from OOXML parts.

    openpyxl's read-only worksheets deliberately omit ``merged_cells``. Reading
    this compact manifest keeps both cell views streaming while preserving merge
    metadata.
    """
    try:
        with ZipFile(io.BytesIO(data)) as archive:
            part_names = set(archive.namelist())
            if _WORKBOOK_PART not in part_names or _WORKBOOK_RELATIONSHIPS_PART not in part_names:
                raise KeyError("workbook parts missing")
            workbook = _xlsx_xml(archive.read(_WORKBOOK_PART))
            relationships = _xlsx_xml(archive.read(_WORKBOOK_RELATIONSHIPS_PART))
            worksheet_targets = {
                relationship.attrib.get("Id"): _safe_xlsx_part(
                    _WORKBOOK_PART,
                    relationship.attrib.get("Target", ""),
                )
                for relationship in relationships.findall(f"{{{_PACKAGE_RELATIONSHIPS_NAMESPACE}}}Relationship")
                if relationship.attrib.get("Type", "").endswith("/worksheet")
                and relationship.attrib.get("TargetMode") != "External"
            }
            merged_ranges = {}
            for sheet in workbook.findall(f".//{{{_WORKBOOK_NAMESPACE}}}sheet"):
                sheet_name = sheet.attrib.get("name")
                relationship_id = sheet.attrib.get(f"{{{_DOCUMENT_RELATIONSHIPS_NAMESPACE}}}id")
                if not sheet_name:
                    continue
                worksheet_part = worksheet_targets.get(relationship_id)
                if worksheet_part not in part_names:
                    merged_ranges[sheet_name] = []
                    continue
                worksheet = _xlsx_xml(archive.read(worksheet_part))
                merged_ranges[sheet_name] = sorted(
                    merge.attrib["ref"]
                    for merge in worksheet.findall(f".//{{{_WORKBOOK_NAMESPACE}}}mergeCell")
                    if merge.attrib.get("ref")
                )
            return merged_ranges
    except (BadZipFile, ET.ParseError, KeyError, OSError) as exc:
        raise SpreadsheetParseError("xlsx_parse_failed", "unable to read XLSX merge metadata") from exc


def _cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return str(value)


def _column_name(column_number: int) -> str:
    value = ""
    while column_number:
        column_number, remainder = divmod(column_number - 1, 26)
        value = chr(65 + remainder) + value
    return value


def _cell_range(row_start: int, row_end: int, column_start: int, column_end: int) -> str:
    start = f"{_column_name(column_start)}{row_start}"
    end = f"{_column_name(column_end)}{row_end}"
    return start if start == end else f"{start}:{end}"


def _append_sheet_rows(document: ParsedDocument, sheet_name: str, rows, merged_ranges: list[str]):
    header_cells = None
    block_index = len(document.text_blocks)
    for row_number, values, cell_provenance, formula_source in rows:
        nonempty_columns = [index for index, value in enumerate(values) if value.strip()]
        if not nonempty_columns:
            continue
        if header_cells is None:
            header_cells = values[:]
        first_column = nonempty_columns[0]
        last_column = nonempty_columns[-1]
        metadata = {
            "sheet_name": sheet_name,
            "headers": header_cells,
            "row_start": row_number,
            "row_end": row_number,
            "column_start": first_column + 1,
            "column_end": last_column + 1,
            "formula_source": formula_source,
            "cell_provenance": cell_provenance,
            "merged_ranges": merged_ranges,
        }
        document.text_blocks.append(
            TextBlock(
                " | ".join(values),
                block_index,
                metadata=metadata,
                block_type="record",
                source_start=row_number,
                source_end=row_number,
            )
        )
        block_index += 1


def parse_xlsx(data: bytes) -> ParsedDocument:
    formula_workbook = None
    cached_workbook = None
    try:
        from openpyxl import load_workbook

        merged_ranges_by_sheet = _xlsx_merged_ranges(data)
        formula_workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=False)
        cached_workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        if formula_workbook is not None:
            formula_workbook.close()
        raise SpreadsheetParseError("xlsx_parse_failed", "unable to read XLSX workbook") from exc

    document = ParsedDocument()
    try:
        for formula_sheet in formula_workbook.worksheets:
            cached_sheet = cached_workbook[formula_sheet.title]
            merged_ranges = merged_ranges_by_sheet.get(formula_sheet.title, [])

            def rows():
                paired_rows = zip(formula_sheet.iter_rows(), cached_sheet.iter_rows())
                for row_number, (formula_cells, cached_cells) in enumerate(paired_rows, start=1):
                    values = []
                    provenance = []
                    has_cached_formula = False
                    has_formula_fallback = False
                    for column_number, (formula_cell, cached_cell) in enumerate(
                        zip(formula_cells, cached_cells),
                        start=1,
                    ):
                        formula = formula_cell.value if getattr(formula_cell, "data_type", None) == "f" else None
                        if formula is not None and cached_cell.value is not None:
                            display_value = cached_cell.value
                            source = "cached"
                            has_cached_formula = True
                        elif formula is not None:
                            display_value = formula
                            source = "formula"
                            has_formula_fallback = True
                        else:
                            display_value = formula_cell.value
                            source = "literal"

                        text = _cell_text(display_value)
                        values.append(text)
                        cell_metadata = {
                            "coordinate": getattr(
                                formula_cell,
                                "coordinate",
                                f"{_column_name(column_number)}{row_number}",
                            ),
                            "column": column_number,
                            "value": text,
                            "source": source,
                        }
                        if formula is not None:
                            cell_metadata["formula"] = _cell_text(formula)
                        provenance.append(cell_metadata)

                    formula_source = "formula" if has_formula_fallback else "cached"
                    if not has_formula_fallback and not has_cached_formula:
                        formula_source = "cached"
                    yield row_number, values, provenance, formula_source

            _append_sheet_rows(document, formula_sheet.title, rows(), merged_ranges)
    finally:
        formula_workbook.close()
        cached_workbook.close()
    return document


def parse_xls(data: bytes) -> ParsedDocument:
    try:
        import xlrd
    except ImportError as exc:
        raise SpreadsheetParseError("xls_parser_unavailable", "xlrd is required to parse XLS files") from exc

    try:
        workbook = xlrd.open_workbook(file_contents=data, on_demand=True, formatting_info=True)
    except Exception as exc:
        raise SpreadsheetParseError("xls_parse_failed", "unable to read XLS workbook") from exc

    document = ParsedDocument()
    try:
        for sheet in workbook.sheets():
            merged_ranges = sorted(
                _cell_range(row_start + 1, row_end, column_start + 1, column_end)
                for row_start, row_end, column_start, column_end in sheet.merged_cells
            )

            def rows():
                for row_number in range(1, sheet.nrows + 1):
                    values = [_cell_text(value) for value in sheet.row_values(row_number - 1)]
                    provenance = [
                        {
                            "coordinate": f"{_column_name(column_number)}{row_number}",
                            "column": column_number,
                            "value": value,
                            "source": "cached",
                        }
                        for column_number, value in enumerate(values, start=1)
                    ]
                    yield row_number, values, provenance, "cached"

            _append_sheet_rows(document, sheet.name, rows(), merged_ranges)
    finally:
        workbook.release_resources()
    return document
